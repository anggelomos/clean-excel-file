from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook

def get_user_input() -> tuple[str, int]:
    """Get input file path and column number from user.

    Returns:
        tuple[str, int]: A tuple containing:
            - input_file (str): Path to the Excel file
            - column_number (int): Column number to process (1-based indexing)
    """
    input_file = input("Ingrese la ruta del archivo Excel: ")
    column_number = int(input("Ingrese el número de columna a limpiar: "))
    return input_file, column_number

def load_sheet(input_file: str) -> Workbook:
    """Load an Excel worksheet from a file.

    Args:
        input_file (str): Path to the Excel file to load.

    Returns:
        Workbook: The loaded workbook.

    Raises:
        FileNotFoundError: If the input file does not exist.
        openpyxl.utils.exceptions.InvalidFileException: If the file is not a valid Excel file.
    """
    workbook = load_workbook(filename=input_file)
    return workbook

def clean_worksheet_data(worksheet: Worksheet, column_number: int) -> Worksheet:
    """Replace commas in cells of a specified column in an Excel worksheet.

    Args:
        worksheet (Worksheet): The Excel worksheet to process.
        column_number (int): The column number to process (1-based indexing).

    Returns:
        Worksheet: The modified worksheet with commas removed from the specified column.

    Raises:
        ValueError: If column_number is not a valid integer or is out of range.
    """
    # Validate column number
    if not isinstance(column_number, int) or column_number < 1 or column_number > worksheet.max_column:
        raise ValueError(f"Column number {column_number} is not valid")
    
    # Clean data
    for row in range(1, worksheet.max_row + 1):
        cell = worksheet.cell(row=row, column=column_number)
        if isinstance(cell.value, str):
            cell.value = cell.value.replace(',', '')
    
    return worksheet

if __name__ == "__main__":
    input_file, column_number = get_user_input()

    workbook = load_workbook(input_file)
    worksheet = workbook.active
    cleaned_worksheet = clean_worksheet_data(worksheet, column_number)

    output_file = input_file.replace('.xlsx', '_limpio.xlsx')
    workbook.save(output_file)
